from os import listdir
from os.path import isfile, join
import main_window_v2
import dialog_window
import traceback
from PyQt5 import QtWidgets
import sys
from parser_happygifts import HappyGifts
from parser_gifts import Gifts
from parser_oasiscatalog import Oasiscatalog
import openpyxl
from datetime import date as g
from os import listdir
from os.path import isfile, join
import time
from PyQt5.QtCore import QThread
import datetime
import os
import json
import pytz
import pickle
from PyQt5.QtWidgets import QMessageBox, QFileDialog
from pathlib import Path
from parser_uzcotton import Uzcotton
from parser_modernit import Modernit


class MainApp(QtWidgets.QMainWindow, main_window_v2.Ui_MainWindow):
    days_weekly = {
        'Понедельник': 0,
        'Вторник': 1,
        'Среда': 2,
        'Четверг': 3,
        'Пятница': 4,
        'Суббота': 5,
        'Воскресенье': 6,
    }

    def __init__(self):
        self.config_save_dir = Path(os.getcwd(), "configs")
        super().__init__()
        self.setupUi(self)
        self.comboBox.addItem('https://happygifts.ru/')
        self.comboBox.addItem('https://gifts.ru/')
        self.comboBox.addItem('https://www.oasiscatalog.com/')
        self.comboBox.addItem('https://uzcotton.ru/')
        self.comboBox.addItem('https://www.modern-it.ru/')
        self.happygifts = HappyGifts(self)
        self.gifts = Gifts(self)
        self.oasiscatalog = Oasiscatalog(self)
        self.uzcotton = Uzcotton(self)
        self.modernit = Modernit(self)
        self.pushButton.clicked.connect(self.update_categorie)
        self.listWidget.itemClicked.connect(self.remove_href)
        self.listWidget_3.itemClicked.connect(self.remove_href)
        self.pushButton_5.clicked.connect(self.update_goods)

        self.pushButton_2.clicked.connect(self.add_in_outlist)
        self.pushButton_8.clicked.connect(self.add_in_outlist_time)

        self.pushButton_3.clicked.connect(self.get_goods)
        self.pushButton_6.clicked.connect(self.add_page)
        self.pushButton_7.clicked.connect(self.add_page_intimelist)
        self.pushButton_4.clicked.connect(self.open_dialog)
        self.pushButton_9.clicked.connect(self.clear_list1)
        self.pushButton_10.clicked.connect(self.clear_list2)

        self.save_config_parse.clicked.connect(
            lambda: self.save_config_parse_to_file(self.config_save_name.toPlainText()))
        self.save_config_parse_timing.clicked.connect(
            lambda: self.save_config_parse_timing_to_file(self.config_save_name.toPlainText()))

        self.update_list_of_configs.clicked.connect(self.read_list_of_configs)
        self.read_list_of_configs()
        self.send_config_to_parse.clicked.connect(
            lambda: self.send_config_to_list(self.combo_configs.currentText(), self.listWidget))
        self.send_config_to_parse_timing.clicked.connect(
            lambda: self.send_config_to_list(self.combo_configs.currentText(), self.listWidget_3))

        self.delete_config.clicked.connect(lambda: self.delete_config_file(self.combo_configs.currentText()))

        self.categories = []
        self.goods = []
        self.timedata = []
        self.choice = []
        self.choice2 = []
        self.sizes = set()
        self.timer = TimerRun(self)
        self.open_json()

    def show_error_message(self, text):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText("Error")
        msg.setInformativeText(text)
        msg.setWindowTitle("Error")
        msg.exec_()

    def are_you_sure(self, text, title="Внимание"):
        buttonReply = QMessageBox.question(self, title, text, QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if buttonReply == QMessageBox.Yes:
            return True
        else:
            return False

    def delete_config_file(self, name_of_config):
        try:
            if (self.are_you_sure(f'вы уверены что хотите удалить конфиг "{name_of_config}"?')):
                os.remove(os.path.join(self.config_save_dir, name_of_config))
                self.read_list_of_configs()
        except Exception as ex:
            traceback.print_exc()

    def read_list_of_configs(self):
        try:
            self.combo_configs.clear()
            configs_files = [f for f in listdir(self.config_save_dir) if isfile(join(self.config_save_dir, f))]
            configs_names = []
            for conf_file in configs_files:
                self.combo_configs.addItem(conf_file)
        except Exception as ex:
            traceback.print_exc()

    def send_config_to_list(self, name_of_config, listWidget):
        try:
            with open(os.path.join(self.config_save_dir, name_of_config), "rb") as f:
                config = pickle.load(f)

            urls = None
            for key, value in config.items():
                urls = value
            for url in urls:
                listWidget.addItem(url)
            self.read_list_of_configs()
        except Exception as ex:
            traceback.print_exc()

    def save_config_parse_to_file(self, name_of_config):
        return self.save_config(name_of_config, self.listWidget)

    def save_config_parse_timing_to_file(self, name_of_config):
        return self.save_config(name_of_config, self.listWidget_3)

    def save_config(self, name_of_config, listWidget):
        try:
            if (name_of_config == ''):
                self.show_error_message("Вы не указали имя файла!!!")
                return
            fileoutput = os.path.join(self.config_save_dir, name_of_config)
            urls = []
            for i in range(0, self.listWidget.count()):
                url = self.listWidget.item(i).text()
                urls.append(url)
            if (len(urls) == 0):
                self.show_error_message("Конфиг пустой, нельзя сохранить конфиг без товара!")
                return
            config = {
                name_of_config: urls
            }
            with open(fileoutput, "wb") as f:
                pickle.dump(config, f)
                self.read_list_of_configs()
            return fileoutput
        except Exception as exc:
            traceback.print_exc()

    def update_categorie(self):
        if self.comboBox.currentText() == 'https://happygifts.ru/':
            self.categories = self.happygifts.parser_category()
        elif self.comboBox.currentText() == 'https://gifts.ru/':
            self.categories = self.gifts.parser_category()
        elif self.comboBox.currentText() == 'https://www.oasiscatalog.com/':
            self.categories = self.oasiscatalog.parser_category()
        elif self.comboBox.currentText() == 'https://uzcotton.ru/':
            self.categories = self.uzcotton.parser_category()
        elif self.comboBox.currentText() == 'https://www.modern-it.ru/':
            self.categories = self.modernit.parser_category()
        self.treeWidget.clear()
        self.listWidget_2.clear()
        for categorie in self.categories:
            item = QtWidgets.QTreeWidgetItem(self.treeWidget)
            item.setText(0, categorie['title'])
            if categorie['subcategories']:
                for subcategorie in categorie['subcategories']:
                    sub_item = QtWidgets.QTreeWidgetItem(item)
                    sub_item.setText(0, subcategorie['title'])

    def update_goods(self):
        self.listWidget_2.clear()
        self.goods = []
        for item in self.treeWidget.selectedItems():
            for categorie in self.categories:
                if categorie['title'] == item.text(0):
                    goods = []
                    if self.comboBox.currentText() == 'https://happygifts.ru/':
                        goods = self.happygifts.parser_goods(categorie['href'])
                    elif self.comboBox.currentText() == 'https://gifts.ru/':
                        goods = self.gifts.parser_goods(categorie['href'])
                    elif self.comboBox.currentText() == 'https://www.oasiscatalog.com/':
                        goods = self.oasiscatalog.parser_goods(categorie['href'])
                    elif self.comboBox.currentText() == 'https://uzcotton.ru/':
                        goods = self.uzcotton.parser_goods(categorie['href'])
                    elif self.comboBox.currentText() == 'https://www.modern-it.ru/':
                        goods = self.modernit.parser_goods(categorie['href'])
                    for good in goods:
                        self.goods.append(good)
                        self.listWidget_2.addItem(good['title'] + ' Артикул:' + good['vendor code'])
                else:
                    if categorie['subcategories']:
                        for subcategorie in categorie['subcategories']:
                            goods = []
                            if subcategorie['title'] == item.text(0):
                                if self.comboBox.currentText() == 'https://happygifts.ru/':
                                    goods = self.happygifts.parser_goods(subcategorie['href'])
                                elif self.comboBox.currentText() == 'https://gifts.ru/':
                                    goods = self.gifts.parser_goods(subcategorie['href'])
                                elif self.comboBox.currentText() == 'https://www.oasiscatalog.com/':
                                    goods = self.oasiscatalog.parser_goods(subcategorie['href'])
                                for good in goods:
                                    self.goods.append(good)
                                    self.listWidget_2.addItem(good['title'] + ' Артикул:' + good['vendor code'])

    def add_in_outlist(self):
        for select_item in self.listWidget_2.selectedItems():
            for good in self.goods:
                item = good['title'] + ' Артикул:' + good['vendor code']
                if item == select_item.text():
                    print(good)
                    if self.comboBox.currentText() == 'https://happygifts.ru/':
                        self.listWidget.addItem('https://happygifts.ru/' + good['href'][1:])
                    elif self.comboBox.currentText() == 'https://gifts.ru/':
                        self.listWidget.addItem('https://gifts.ru/' + good['href'][1:])
                    elif self.comboBox.currentText() == 'https://www.oasiscatalog.com/':
                        self.listWidget.addItem('https://www.oasiscatalog.com/' + good['href'][1:])
                    elif self.comboBox.currentText() == 'https://www.modern-it.ru/':
                        self.listWidget.addItem('https://www.modern-it.ru/' + good['href'][1:])
                    elif self.comboBox.currentText() == 'https://uzcotton.ru/':
                        self.listWidget.addItem('https://uzcotton.ru/' + good['href'][1:])
                    break

    def add_in_outlist_time(self):
        for select_item in self.listWidget_2.selectedItems():
            for good in self.goods:
                item = good['title'] + ' Артикул:' + good['vendor code']
                if item == select_item.text():
                    if self.comboBox.currentText() == 'https://happygifts.ru/':
                        self.listWidget_3.addItem('https://happygifts.ru/' + good['href'][1:])
                    elif self.comboBox.currentText() == 'https://gifts.ru/':
                        self.listWidget_3.addItem('https://gifts.ru/' + good['href'][1:])
                    elif self.comboBox.currentText() == 'https://www.oasiscatalog.com/':
                        self.listWidget_3.addItem('https://www.oasiscatalog.com/' + good['href'][1:])
                    elif self.comboBox.currentText() == 'https://www.modern-it.ru/':
                        self.listWidget_3.addItem('https://www.modern-it.ru/' + good['href'][1:])
                    elif self.comboBox.currentText() == 'https://uzcotton.ru/':
                        self.listWidget_3.addItem('https://uzcotton.ru/' + good['href'][1:])

    def get_goods(self):
        filename = self.saveFileDialog()
        if not filename:
            return
        goods = []
        for i in range(0, self.listWidget.count()):
            main_page = self.listWidget.item(i).text()
            if self.happygifts.main_page in main_page:
                driver = self.happygifts
            elif self.gifts.main_page in main_page:
                driver = self.gifts
            elif self.oasiscatalog.main_page in main_page:
                driver = self.oasiscatalog
            elif self.modernit.main_page in main_page:
                driver = self.modernit
            elif self.uzcotton.main_page in main_page:
                driver = self.uzcotton
            goods.append(driver.parser_good(main_page))

        return self.parsing(goods, filename)

    def get_goods_timing(self):
        filename = self.saveFileDialog()
        if not filename:
            return
        goods = []
        for i in range(0, self.listWidget_3.count()):
            main_page = self.listWidget_3.item(i).text()
            if self.happygifts.main_page in main_page:
                driver = self.happygifts
            elif self.gifts.main_page in main_page:
                driver = self.gifts
            elif self.oasiscatalog.main_page in main_page:
                driver = self.oasiscatalog
            elif self.modernit.main_page in main_page:
                driver = self.modernit
            elif self.uzcotton.main_page in main_page:
                driver = self.uzcotton
            goods.append(driver.parser_good(main_page))
        return self.parsing(goods, filename)

    def parsing(self, goods, fileoutput):
        try:
            wb = openpyxl.Workbook()
            sheet = wb.create_sheet(index=0, title='Sheet 1')
            titles = ['Артикул', 'Раздел', 'Наименование', 'Ссылка', 'Цвет', 'Отметка', 'Цена', ' ']
            titles.extend(self.get_sort_size_list())
            titles.extend(['Описание', 'Материалы'])
            for col in range(1, len(titles) + 1):
                sheet.cell(1, col, titles[col - 1])
            count = 0
            row = 2
            for i in range(len(goods)):
                if goods[count] is not None:
                    for j in range(len(goods[count]['page'])):
                        index = titles.index('Описание')
                        sheet.cell(row, 1, goods[count]['vendor_code'][j])
                        sheet.cell(row, 2, goods[count]['section'])
                        sheet.cell(row, 3, goods[count]['name'])
                        sheet.cell(row, 4, goods[count]['page'][j])
                        sheet.cell(row, 5, goods[count]['color'][j])
                        sheet.cell(row, 6, goods[count]['marks'][j])
                        sheet.cell(row, 7, goods[count]['price'][j])
                        sheet.cell(row, index + 1, goods[count]['descriptions'][j])
                        sheet.cell(row, index + 2, goods[count]['materials'])
                        sizes = [(key, val) for key, val in goods[count]['sizes'][j].items()]
                        for s, amount in sizes:
                            sheet.cell(row, titles.index(s) + 1, amount)
                        row += 1
                count += 1
            wb.save(fileoutput)
            print('Succsessfully')
        except Exception as ex:
            traceback.print_exc()
            print(ex)
            print(goods)
        finally:
            self.sizes.clear()

    def get_sort_size_list(self):
        sizes = ['XS', 'S', 'M',  'L', 'XL', 'XXL', '3XL', '4XL', '5XL', '6XL']
        lst = []
        for size in sizes:
            if size in self.sizes:
                lst.append(size)
        for size in self.sizes:
            if size not in lst:
                lst.append(size)
        return lst

    def get_size(self, size):
        sizes = ['XS', '3XL', '4XL', '5XL', '6XL', 'XXL', 'XL', 'S', 'M', 'L']
        for i in sizes:
            if i in size:
                size = i
                break
        self.sizes.add(size)
        return size

    def add_page(self):
        item = self.lineEdit.text()
        self.listWidget.addItem(item)

    def add_page_intimelist(self):
        item = self.lineEdit_2.text()
        self.listWidget_3.addItem(item)

    def remove_href(self):
        for item in self.listWidget.selectedItems():
            self.listWidget.takeItem(self.listWidget.row(item))
        for item in self.listWidget_3.selectedItems():
            self.listWidget_3.takeItem(self.listWidget_3.row(item))

    def clear_list1(self):
        self.listWidget.clear()

    def clear_list2(self):
        self.listWidget_3.clear()

    def open_dialog(self):
        if self.pushButton_4.isChecked():
            dialog = Dialog()
            dialog.exec_()
            if dialog.accepted:
                get_time = dialog.time.split(':')

                self.timedata = [dialog.day,
                                 datetime.time(hour=int(get_time[0]),
                                               minute=int(get_time[1][:2])),
                                 int(dialog.count_pars),
                                 0, dialog.selected_timezone]
                if self.timedata[2] > 0:
                    self.timer.timedata = self.timedata
                    self.timer.start()
                    self.pushButton_4.setText('Остановить парсинг')
                    self.save_json()
                else:
                    self.pushButton_4.setChecked(False)
                    self.pushButton_4.setText('Парсинг')
            else:
                self.pushButton_4.setChecked(False)
                self.pushButton_4.setText('Парсинг')
        else:
            self.pushButton_4.setText('Парсинг')
            self.timer.timedata[2] = 0

    def parsing_time2(self):
        if self.listWidget_3.count() > 0:
            self.get_goods_timing()

    def save_json(self):
        try:
            timedata = [self.timer.timedata[0],
                        str(self.timer.timedata[1]),
                        self.timer.timedata[2],
                        self.timer.timedata[3],
                        self.timer.timedata[4]]
            to_json = {'timedata': timedata,
                       'pages': [self.listWidget_3.item(i).text() for i in range(0, self.listWidget_3.count())]}
            with open('timedata.json', 'w') as f:
                json.dump(to_json, f)
        except Exception as ex:
            traceback.print_exc()
            print(ex)

    def open_json(self):
        if os.path.exists('timedata.json'):
            with open('timedata.json') as f:
                timedata_file = json.load(f)
                if timedata_file['timedata'][2] > 0:
                    for page in timedata_file['pages']:
                        self.listWidget_3.addItem(page)
                    get_time = timedata_file['timedata'][1].split(':')
                    self.timer.timedata = [timedata_file['timedata'][0],
                                           datetime.time(hour=int(get_time[0]), minute=int(get_time[1])),
                                           timedata_file['timedata'][2],
                                           timedata_file['timedata'][3],
                                           timedata_file['timedata'][4]]
                    self.timer.start()
                    self.pushButton_4.setText('Остановиь парсинг')
                    self.pushButton_4.setChecked(True)

    def saveFileDialog(self):
        fileName, _ = QFileDialog.getSaveFileName(self, "Save as...", 'data.xlsx', "Excel (*.xlsx)")
        if fileName:
            return fileName


class Dialog(QtWidgets.QDialog, dialog_window.Ui_Dialog):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.buttonBox.accepted.connect(self.accept)
        self.buttonBox.rejected.connect(self.cancel)
        self.comboBox.addItems(pytz.all_timezones)
        self.comboBox.setCurrentText('Europe/Moscow')
        self.accepted = False
        self.day = []
        self.time = None
        self.count_pars = None
        self.selected_timezone = None

    def accept(self):
        check_boxes = [self.checkBox.isChecked(),
                       self.checkBox_2.isChecked(),
                       self.checkBox_3.isChecked(),
                       self.checkBox_4.isChecked(),
                       self.checkBox_5.isChecked(),
                       self.checkBox_6.isChecked(),
                       self.checkBox_7.isChecked()]
        for i in range(0, 7):
            if check_boxes[i]:
                self.day.append(i)
        print(self.day)
        self.time = self.timeEdit.text()
        self.count_pars = self.spinBox_2.text()
        self.selected_timezone = self.comboBox.currentText()
        self.accepted = True
        self.close()

    def cancel(self):
        self.accepted = False
        self.close()


class TimerRun(QThread):
    def __init__(self, mainWindow):
        super().__init__()
        self.value = 0
        self.mainWindow = mainWindow
        self.timedata = []
        self.parsed = []

    def run(self):
        while True:
            if self.timedata[2] > 0:
                time.sleep(1)
                now = datetime.datetime.now(tz=pytz.timezone(self.timedata[4]))
                if now.weekday() in self.timedata[0] \
                        and now.time().minute == self.timedata[1].minute \
                        and now.time().hour == self.timedata[1].hour:
                    if [now.day, now.time().hour, now.time().minute] not in self.parsed:
                        print('[INFO] PARSING...')
                        self.mainWindow.parsing_time2()
                        self.parsed.append([now.day, now.time().hour, now.time().minute])
                        print('[INFO] complete')
                        self.timedata[2] -= 1
                        self.timedata[3] += 1
                        self.mainWindow.save_json()
            else:
                self.mainWindow.pushButton_4.setChecked(False)
                self.mainWindow.pushButton_4.setText('Парсинг')
                self.mainWindow.save_json()
                break


def except_hook(cls, exception, traceback):
    sys.__excepthook__(cls, exception, traceback)


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    window = MainApp()
    sys.excepthook = except_hook
    window.show()
    app.exec_()
